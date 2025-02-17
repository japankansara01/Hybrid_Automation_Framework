import openpyxl

def get_rowcount(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return(sheet.max_row)

def get_columncount(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return(sheet.max_column)

def read_data(file,sheetname,rownum,columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.cell(rownum,columnno).value

def write_data(file,sheetname,rownum,columnno,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(rownum,columnno).value = data
    workbook.save(file)

# def fillGreenColor(file,sheetName,rownum,columnno):
#     workbook = openpyxl.load_workbook(file)
#     sheet = workbook[sheetName]
#     greenFill = PatternFill(start_color='60b212',end_color='60b212',fill_type='solid')
#     sheet.cell(rownum, columnno).fill=greenFill
#     workbook.save(file)
#
# def fillRedColor(file,sheetName,rownum,columnno):
#     workbook = openpyxl.load_workbook(file)
#     sheet = workbook[sheetName]
#     redFill = PatternFill(start_color='ff0000',end_color='ff0000',fill_type='solid')
#     sheet.cell(rownum, columnno).fill=redFill
#     workbook.save(file)